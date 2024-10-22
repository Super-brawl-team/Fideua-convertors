import os
import sys
from PIL import Image
from pdf2image import convert_from_path
from docx import Document
from openpyxl import load_workbook
import numpy as np

def readRawFileData(inputPath):
    try:
        with open(inputPath, 'rb') as f:
            return f.read()
    except Exception as e:
        print(f"Error reading file: {e}")
        return None

def convertImageToPng(inputPath, outputPath):
    try:
        img = Image.open(inputPath)
        img.save(outputPath, 'PNG')
        print(f"Image saved as PNG: {outputPath}")
    except Exception as e:
        print(f"Error converting image: {e}")

def convertPdfToPng(inputPath, outputDir):
    try:
        pages = convert_from_path(inputPath)
        for i, page in enumerate(pages):
            outputPath = os.path.join(outputDir, f"page_{i + 1}.png")
            page.save(outputPath, 'PNG')
            print(f"PDF page saved as PNG: {outputPath}")
    except Exception as e:
        print(f"Error converting PDF: {e}")

def convertDocxToPng(inputPath, outputDir):
    try:
        doc = Document(inputPath)
        for i, paragraph in enumerate(doc.paragraphs):
            img = Image.new('RGB', (800, 400), color = (255, 255, 255))
            d = ImageDraw.Draw(img)
            d.text((10,10), paragraph.text, fill=(0,0,0))
            outputPath = os.path.join(outputDir, f"paragraph_{i + 1}.png")
            img.save(outputPath, 'PNG')
            print(f"DOCX paragraph saved as PNG: {outputPath}")
    except Exception as e:
        print(f"Error converting DOCX: {e}")

def convertXlsxToPng(inputPath, outputDir):
    try:
        wb = load_workbook(inputPath)
        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            data = "\n".join([", ".join([str(cell.value) for cell in row]) for row in sheet.iter_rows()])

            img = Image.new('RGB', (800, 800), color = (255, 255, 255))
            d = ImageDraw.Draw(img)
            d.text((10, 10), data, fill=(0, 0, 0))
            outputPath = os.path.join(outputDir, f"{sheetname}.png")
            img.save(outputPath, 'PNG')
            print(f"XLSX sheet saved as PNG: {outputPath}")
    except Exception as e:
        print(f"Error converting XLSX: {e}")

def convertUnsupportedFileToPng(inputPath, outputPath):
    rawData = readRawFileData(inputPath)
    if rawData is None:
        return

    fileSize = len(rawData)
    width = int(fileSize ** 0.5)
    height = fileSize // width

    pixelData = np.frombuffer(rawData[:width * height], dtype=np.uint8)
    pixelData = pixelData.reshape((height, width))

    img = Image.fromarray(pixelData, mode='L')

    img.save(outputPath)
    print(f"PNG saved as: {outputPath}")

def fideuaConvertToPng(inputPath):
    fileExtension = inputPath.split('.')[-1].lower()
    outputDir = os.path.dirname(inputPath)

    if fileExtension in ['jpg', 'jpeg', 'bmp', 'gif', 'tiff', 'webp']:
        outputPath = os.path.join(outputDir, os.path.splitext(os.path.basename(inputPath))[0] + '.png')
        convertImageToPng(inputPath, outputPath)
    elif fileExtension == 'pdf':
        convertPdfToPng(inputPath, outputDir)
    elif fileExtension == 'docx':
        convertDocxToPng(inputPath, outputDir)
    elif fileExtension == 'xlsx':
        convertXlsxToPng(inputPath, outputDir)
    else:
        outputPath = os.path.join(outputDir, os.path.splitext(os.path.basename(inputPath))[0] + 'Troll.png')
        convertUnsupportedFileToPng(inputPath, outputPath)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python convertAnyToPng.py <file>")
    else:
        inputFile = sys.argv[1]
        fideuaConvertToPng(inputFile)
